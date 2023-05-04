'''
dNTP_plate_preparation
Fill the dNTP_plate_preparation excel file
Script: Ludivine Profit
Last update: 2020/12/03
'''

path = r'C:\Users\Hamilton\Desktop\HAMILTON_control\plate_preparation\dNTP plat design\dNTP_plate_design.xlsx'
outputPath=r'C:\Users\Hamilton\Desktop\HAMILTON_control\plate_preparation\dNTP plat design\dNTP_plate_preparation.xlsx'
import openpyxl
import xlrd

def getExcelSheet(path):
    """
    Returns the first sheet of the excel at the path
    :param path:
    :return:
    """
    wb = xlrd.open_workbook(path)
    dNTP_sheet = wb.sheet_by_index(0)
    return dNTP_sheet

def getPlateOrder(dNTP_sheet):
    """
    Returns the plate order from dNTP_sheet
    :param dNTP_sheet
    :return plateOrder (arr of int)
    """
    plateOrder =[]
    plateOrder_indexes = findIndexes("Plate order", dNTP_sheet)
    cycle = 1
    plate = 1

    while plate != "":
        plate = dNTP_sheet.cell_value(plateOrder_indexes[0], plateOrder_indexes[1] + cycle)
        cycle = cycle + 1
        if plate != "":
            plateOrder.append(int(plate))
    return plateOrder

def getSequences(dNTP_sheet):
    """
    Returns an array of 96 sequences from dNTP_sheet
    :param dNTP_sheet
    :return sequences (arr of 96 str)
    """
    sequences=[]

    sequenceLayout_indexes = findIndexes("Sequence layout", dNTP_sheet)

    for col in range(1,13):
        for row in range(1,9):
            sequence=str(dNTP_sheet.cell(sequenceLayout_indexes[0]+row,sequenceLayout_indexes[1]+col).value)
            sequences.append(sequence)

    return sequences

def getConcentrations(dNTP_sheet):
    """
    Returns an array of 96 dNTP final concentrations from dNTP_sheet
    :param dNTP_sheet
    :return concentrations (arr of 96 flt)
    """
    concentrations=[]

    concentration_indexes = findIndexes("Concentrations", dNTP_sheet)

    for col in range(1,13):
        for row in range(1,9):
            concentration=dNTP_sheet.cell(concentration_indexes[0]+row,concentration_indexes[1]+col).value
            concentrations.append(concentration)

    return concentrations

def getVolumes(dNTP_sheet):
    """
    Returns an array of 4 final volumes (1/dNTP plate) from dNTP_sheet
    :param dNTP_sheet
    :return volumes (arr of 4 flt)
    """
    volumes = []

    volumes_indexes = findIndexes("Volume/well (µL)", dNTP_sheet)

    for plate in range(1,5):
            volume = dNTP_sheet.cell(volumes_indexes[0],volumes_indexes[1]+plate).value
            volumes.append(volume)
    return volumes

def getStockConcetrations(dNTP_sheet):
    """
    Returns an array of 4 stock concentrations (1/dNTP) from dNTP_sheet
    :param dNTP_sheet
    :return stockConcentrations (arr of 4 flt)
    """
    stockConcentrations = []

    dNTPstockConcentrations_indexes = findIndexes("[dNTP] stock", dNTP_sheet)

    for dNTP in range(1,5):
            stockConcentration = dNTP_sheet.cell(dNTPstockConcentrations_indexes[0],dNTPstockConcentrations_indexes[1]+dNTP).value
            stockConcentrations.append(stockConcentration)
    return stockConcentrations

def getPlates(sequences,plateOrder,concentrations):
    """
    Returns an array of 96 wells for 4 nuc plates for 4 premix plates
    :param sequences (arr of 96 str)
    :param plateOrder (arr of int)
    :param concentrations (arr of 96 flt)
    :return plates (arr)
    """
    #plates = array of 96 wells for 4 nuc plates for 4 premix plates
    #x = number of well
    x = 96
    plates=[[[0 for well in range (x)]for nucPlates in range (4)]for premixPlates in range (4)]
    well = 0
    for sequence in sequences:
        i = 0
        for nuc in sequence:
                if nuc == "A":
                    plates[plateOrder[i]-1][0][well] = concentrations[well]
                if nuc == "C":
                    plates[plateOrder[i]-1][1][well]= concentrations[well]
                if nuc == "G":
                    plates[plateOrder[i]-1][2][well]= concentrations[well]
                if nuc == "T":
                    plates[plateOrder[i]-1][3][well]= concentrations[well]
                i = i + 1
        well = well + 1
    return plates

def writeExcel(volumes, stockConcentrations, plates):
    """
    Fill dNTP_plate_preparation Excel file with stock and final dNTP concentrations and final volumes of each dNTP plate
    :param volumes (arr of 96 flt)
    :param stockConcentrations (arr of 4 flt)
    :param plates (arr)
    :return
    """
    #open plate preparation excel file to write
    Plate_preparation_excel = openpyxl.load_workbook(outputPath)
    Plates_sheet = Plate_preparation_excel["Plates"]

    offsetRow=12
    offsetCol=2

    for premixPlate in range (4):
        #find column index of the premix plate
        plateStr = "Plate " + str(premixPlate+1)
        for col in range(1,Plates_sheet.max_column):
            if Plates_sheet.cell(1, col).value == plateStr:
                premixPlate_col=col

        #write final volume per premix plate
        for row in range(1,Plates_sheet.max_row):
            if Plates_sheet.cell(row, premixPlate_col).value == "Final volume":
                finalVolume_indexes=[row,premixPlate_col]
                Plates_sheet.cell(finalVolume_indexes[0],finalVolume_indexes[1]+1).value = volumes[premixPlate]

        #write stock concentration and final concentrations per nucleotide
        for nucPlate in range (4):
            Plates_sheet.cell(nucPlate*14+9,premixPlate_col+1).value = stockConcentrations[nucPlate]
            well = 0
            for col in range(12):
                for row in range(8):
                    Plates_sheet.cell(offsetRow + row, premixPlate_col + offsetCol + col).value = plates[premixPlate][nucPlate][well]
                    well = well + 1
            offsetRow = offsetRow + 14
        offsetRow = 12

    Plate_preparation_excel.save(outputPath)

def findIndexes(eltToFind,dNTP_sheet):
    """
    Find in an excel sheet the indexes of the cell containing the string 'eltToFind'
    :param eltToFind: a string to find in the excel file
    :param dNTP_sheet:
    :return:
    """

    for row in range(dNTP_sheet.nrows):
        for col in range(dNTP_sheet.ncols):
            if dNTP_sheet.cell_value(row,col)==eltToFind:
                return (row,col)

dNTP_sheet=getExcelSheet(path)

plateOrder=getPlateOrder(dNTP_sheet)
print("plateOrder = ")
print(plateOrder)

sequences=getSequences(dNTP_sheet)
print("sequences = ")
print(sequences)

concentrations=getConcentrations(dNTP_sheet)

print("volumes =")
volumes=getVolumes(dNTP_sheet)
print(volumes)

print("[dNTP] stock =")
stockConcentrations=getStockConcetrations(dNTP_sheet)
print(stockConcentrations)

plates=getPlates(sequences,plateOrder,concentrations)

writeExcel(volumes, stockConcentrations, plates)
