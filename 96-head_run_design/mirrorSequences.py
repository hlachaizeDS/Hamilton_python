'''
mirrorSequences
Generate 24 mirror sequences with 4 nuc from a "mother sequence"
Script: Ludivine Profit
Last update: 2020/12/03
'''

path = r'C:\Users\Hamilton\Desktop\HAMILTON_control\plate_preparation\dNTP plat design\dNTP_plate_design.xlsx'
outputPath = r'C:\Users\Hamilton\Desktop\HAMILTON_control\plate_preparation\dNTP plat design\Mirror_sequences.xlsx'
import openpyxl

def getMotherSequences(path):
    """
    Returns the mother sequence from dNTP_plate_design file
    :param path
    :return sequence (str): mother sequence
    """
    dNTP_excel = openpyxl.load_workbook(path)
    dNTP_sheet = dNTP_excel["dNTP"]
    motherSequenceIndexes = []
    nrow = dNTP_sheet.max_row + 1
    ncol = dNTP_sheet.max_column + 1

    for row in range(1, nrow):
        for col in range(1, ncol):
            value = dNTP_sheet.cell(row,col).value
            if value == "Mother sequence":
                motherSequenceIndexes = [row,col]
    sequence = dNTP_sheet.cell(motherSequenceIndexes[0], motherSequenceIndexes[1]+1).value
    return sequence

def plateOrderFromSequence(plateID,sequence):
    """
    Returns plate order from mother sequence
    :param plateID (arr of 4 int): corresponding number of the plates for [A,C,G,T]
    :param sequence (str): mother sequence
    :return plateOrder (arr of int)
    """
    #plateID being the corresponding number of the plates for [A,C,G,T]
    plateOrder=[]
    for nuc in sequence:
        if nuc =="A":
            plateOrder.append(plateID[0])
        if nuc =="C":
            plateOrder.append(plateID[1])
        if nuc == "G":
            plateOrder.append(plateID[2])
        if nuc == "T":
            plateOrder.append(plateID[3])
    return plateOrder

def sequenceFromPlateOrder(plateID,plateOrder):
    """
    Returns 1 sequence from plate IDs and plate order
    :param plateID (arr of 4 int): corresponding number of the plates for [A,C,G,T]
    :param plateOrder (arr of int)
    :return sequence (str)
    """
    sequence = ""
    for number in plateOrder:
        index=plateID.index(number)+1
        if index == 1:
            sequence = sequence + "A"
        if index == 2:
            sequence = sequence + "C"
        if index == 3:
            sequence = sequence + "G"
        if index == 4:
            sequence = sequence + "T"


    return sequence

def allMirrorSequences(sequence):
    """
    Returns 24 mirror sequences with the 4 bases from 1 mother sequence
    :param sequence (str): mother sequence
    :return mirrorSequences (arr of str)
    """
    initialPlateID=[1,2,3,4]
    plateOrder=plateOrderFromSequence(initialPlateID,sequence)
    mirrorSequences=[]
    print("plateOrder = ")
    print(plateOrder)
    seqNb=1

    for Aindex in list(range(1,4+1)):
        Clist=list(range(1,4+1))
        Clist.remove(Aindex)
        for Cindex in Clist:
            Glist=list(range(1,4+1))
            Glist.remove(Aindex)
            Glist.remove(Cindex)
            for Gindex in Glist:
                Tlist = list(range(1, 4 + 1))
                Tlist.remove(Aindex)
                Tlist.remove(Cindex)
                Tlist.remove(Gindex)
                for Tindex in Tlist:
                    print("-------------------------------")
                    print("Mirror Sequence #" + str(seqNb) )
                    print("A = " + str(Aindex) +", C = " + str(Cindex) +", G = " + str(Gindex) +", T = " + str(Tindex))
                    plateID=[Aindex,Cindex,Gindex,Tindex]
                    print(plateID)
                    print(sequenceFromPlateOrder(plateID,plateOrder))
                    mirrorSequences.append(sequenceFromPlateOrder(plateID,plateOrder))
                    seqNb=seqNb+1
    return mirrorSequences


def writeMirrorSequencesExcel(plateOrder,mirrorSequences):
    """
    Write 24 mirror sequences and plate order in mirror_sequences file
    :param plateOrder (arr of int)
    :param mirrorSequences (arr of str): array of 24 mirror sequences with the 4 bases
    :return 
    """
    Mirror_excel = openpyxl.load_workbook(outputPath)
    Mirror_sheet = Mirror_excel["Mirror"]

    # write plateOrder in Excel output
    plateOrderIndexes = []
    nrow = Mirror_sheet.max_row + 1
    ncol = Mirror_sheet.max_column + 1

    for row in range(1,nrow):
        for col in range(1,ncol):
            value = Mirror_sheet.cell(row, col).value
            if value == "Plate order":
                plateOrderIndexes = [row, col]


    # erase previous platOrder
    for col in range(1, ncol):
        Mirror_sheet.cell(plateOrderIndexes[0], plateOrderIndexes[1] + col).value = ""

    # write plateOrder in Excel output
    col = 1
    for plateNbr in plateOrder:
        Mirror_sheet.cell(plateOrderIndexes[0], plateOrderIndexes[1] + col).value = plateNbr
        col = col + 1

    #Write mirror sequences
    mirrorSequencesIndexes = []
    for row in range(1, 100):
        for col in range(1, 100):
            value = Mirror_sheet.cell(row, col).value
            if value == "Mirror sequences":
                mirrorSequencesIndexes = [row, col]
    row = 1
    for seq in mirrorSequences:
        Mirror_sheet.cell(mirrorSequencesIndexes[0] + row, mirrorSequencesIndexes[1] + 1).value = seq
        row = row + 1

    Mirror_excel.save(outputPath)

plateID=[1,2,3,4]
sequence=getMotherSequences(path)
print(sequence)
plateOrder=plateOrderFromSequence(plateID,sequence)
mirrorSequences = allMirrorSequences(sequence)
writeMirrorSequencesExcel(plateOrder,mirrorSequences)

